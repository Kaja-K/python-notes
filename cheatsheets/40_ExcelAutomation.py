import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import numpy as np

# ════════════════════════════════════════════
#  OPENPYXL — create / edit .xlsx without Excel installed
# ════════════════════════════════════════════

# ── CREATE & SAVE ────────────────────────────────────────────
wb = Workbook()
ws = wb.active                          # get default sheet
ws.title = 'Report'                     # rename sheet

wb.save('output.xlsx')                  # save to file
wb.close()

# open existing file
wb = load_workbook('data.xlsx')
wb = load_workbook('data.xlsx', read_only=True)   # faster, no editing
wb = load_workbook('data.xlsx', data_only=True)   # read values not formulas

# ── SHEETS ───────────────────────────────────────────────────
wb.sheetnames                           # ['Sheet', 'Data', 'Summary']
ws  = wb['Sheet']                       # get sheet by name
ws  = wb.active                         # currently active sheet

ws2 = wb.create_sheet('Summary')        # add new sheet at end
ws3 = wb.create_sheet('Intro', 0)       # insert at position 0 (first)
wb.remove(ws2)                          # delete a sheet
wb.copy_worksheet(ws)                   # duplicate a sheet

ws.sheet_properties.tabColor = '1072BA' # set tab colour (hex)

# ── READING & WRITING CELLS ──────────────────────────────────
ws['A1'] = 'Station'                    # write by cell address
ws['B1'] = 'Load %'
ws['A2'] = 'ST_001'
ws['B2'] = 85.5

ws['A1'].value                          # read a value
ws.cell(row=2, column=1).value          # read by row/col numbers
ws.cell(row=2, column=1, value='ST_002') # write by row/col

# write a formula
ws['C1'] = '=SUM(B2:B100)'
ws['D1'] = '=AVERAGE(B2:B100)'
ws['E1'] = '=VLOOKUP(C2, Sheet2!A1:B10, 2, FALSE)'

# ── READING RANGES ───────────────────────────────────────────
ws['A1':'C10']                          # range of cells
for row in ws.iter_rows(min_row=2, max_row=10, min_col=1, max_col=3):
    for cell in row:
        print(cell.value)

for row in ws.iter_rows(min_row=1, values_only=True):
    print(row)                          # tuple of values per row

ws.max_row                              # last row with data
ws.max_column                          # last column with data

# ── WRITING MULTIPLE ROWS ────────────────────────────────────
data = [
    ['ST_001', 'Warsaw', 85.5],
    ['ST_002', 'Krakow', 42.0],
    ['ST_003', 'Wroclaw', 91.3],
]
for row in data:
    ws.append(row)                      # append a row at the bottom

# ── COLUMN & ROW DIMENSIONS ──────────────────────────────────
ws.column_dimensions['A'].width = 15
ws.row_dimensions[1].height = 25
ws.column_dimensions['B'].hidden = True # hide a column

# freeze panes — keep header visible when scrolling
ws.freeze_panes = 'A2'                  # freeze row 1
ws.freeze_panes = 'B1'                  # freeze column A

# ── STYLING ──────────────────────────────────────────────────
# font
ws['A1'].font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')

# fill (background colour)
ws['A1'].fill = PatternFill(fill_type='solid', fgColor='1072BA')  # blue

# alignment
ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# border
thin = Side(style='thin', color='CCCCCC')
ws['A1'].border = Border(top=thin, bottom=thin, left=thin, right=thin)

# number format
ws['B2'].number_format = '0.00'         # 2 decimal places
ws['B3'].number_format = '#,##0'        # thousands separator
ws['B4'].number_format = '0.0%'         # percentage
ws['C1'].number_format = 'YYYY-MM-DD'   # date

# ── STYLE A RANGE ────────────────────────────────────────────
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(fill_type='solid', fgColor='1F4E79')

for cell in ws[1]:                      # all cells in row 1
    cell.font  = header_font
    cell.fill  = header_fill
    cell.alignment = Alignment(horizontal='center')

# ── CONDITIONAL FORMATTING ───────────────────────────────────
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# colour scale — green → yellow → red
ws.conditional_formatting.add('B2:B100', ColorScaleRule(
    start_type='min', start_color='63BE7B',
    mid_type='percentile', mid_value=50, mid_color='FFEB84',
    end_type='max', end_color='F8696B'
))

# highlight cells > 80
ws.conditional_formatting.add('B2:B100', CellIsRule(
    operator='greaterThan', formula=['80'],
    fill=PatternFill(fill_type='solid', fgColor='FF0000')
))

# ── DATA VALIDATION ──────────────────────────────────────────
# dropdown list
dv = DataValidation(type='list', formula1='"Active,Inactive,Maintenance"', allow_blank=True)
ws.add_data_validation(dv)
dv.add('C2:C100')                       # apply to range

# number range validation
dv2 = DataValidation(type='whole', operator='between', formula1=0, formula2=100)
dv2.error = 'Value must be between 0 and 100'
ws.add_data_validation(dv2)
dv2.add('B2:B100')

# ── CHARTS ───────────────────────────────────────────────────
# bar chart
chart = BarChart()
chart.title  = 'Load by Station'
chart.y_axis.title = 'Load %'
chart.x_axis.title = 'Station'

data_ref    = Reference(ws, min_col=2, min_row=1, max_row=10)  # data series
cats_ref    = Reference(ws, min_col=1, min_row=2, max_row=10)  # categories
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws.add_chart(chart, 'E2')               # place chart at cell E2

# line chart
line = LineChart()
line.title = 'Load over Time'
line.add_data(Reference(ws, min_col=2, min_row=1, max_row=20), titles_from_data=True)
ws.add_chart(line, 'E20')

# pie chart
pie = PieChart()
pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=6), titles_from_data=True)
pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))
ws.add_chart(pie, 'M2')

# ── UTILITY FUNCTIONS ────────────────────────────────────────
get_column_letter(3)            # 'C'
get_column_letter(26)           # 'Z'
get_column_letter(27)           # 'AA'
column_index_from_string('C')   # 3
column_index_from_string('AA')  # 27

# ── PROTECTION ───────────────────────────────────────────────
ws.protection.set_password('secret')   # protect sheet from editing
wb.security.lockStructure = True        # prevent adding/deleting sheets

# ════════════════════════════════════════════
#  PANDAS + OPENPYXL — best of both worlds
# ════════════════════════════════════════════

np.random.seed(42)
df = pd.DataFrame({
    'station_id':  [f'ST_{str(i).zfill(3)}' for i in range(1, 101)],
    'city':        np.random.choice(['Warsaw','Krakow','Wroclaw','Gdansk','Poznan'], 100),
    'device_type': np.random.choice(['Router','Switch','BTS','Repeater'], 100),
    'status':      np.random.choice(['Active','Inactive','Maintenance'], 100, p=[0.7,0.2,0.1]),
    'load_pct':    np.round(np.random.uniform(5, 100, 100), 1),
    'uptime_days': np.random.randint(1, 730, 100),
    'incidents':   np.random.poisson(lam=1.5, size=100),
    'temperature': np.round(np.random.normal(45, 10, 100), 1),
})

# write DataFrame to Excel
df.to_excel('report.xlsx', index=False, sheet_name='Data')

# write multiple sheets
with pd.ExcelWriter('multi_sheet.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Raw Data', index=False)
    df.groupby('city')['load_pct'].mean().to_excel(writer, sheet_name='Summary')
    df[df['load_pct'] > 80].to_excel(writer, sheet_name='High Load', index=False)

# write DataFrame then add formatting on top
with pd.ExcelWriter('styled.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Report', index=False)
    wb_out = writer.book
    ws_out = writer.sheets['Report']

    # style the header row
    for cell in ws_out[1]:
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = PatternFill(fill_type='solid', fgColor='1F4E79')
        cell.alignment = Alignment(horizontal='center')

    # auto-width columns
    for col in ws_out.columns:
        max_len = max(len(str(c.value or '')) for c in col)
        ws_out.column_dimensions[col[0].column_letter].width = max_len + 4

# read Excel into DataFrame
df_in = pd.read_excel('report.xlsx', sheet_name='Data')
df_in = pd.read_excel('multi_sheet.xlsx', sheet_name='Summary')
df_in = pd.read_excel('multi_sheet.xlsx', sheet_name=None)  # dict of all sheets

# ── DATA CLEANING + EXPORT PATTERN ──────────────────────────
def export_report(df: pd.DataFrame, path: str) -> None:
    """clean data, build summary, export to styled Excel"""
    pivot = df.groupby('city').agg(
        stations   = ('station_id', 'count'),
        avg_load   = ('load_pct', 'mean'),
        max_load   = ('load_pct', 'max'),
        incidents  = ('incidents', 'sum')
    ).round(1).reset_index()

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer,    sheet_name='Raw',     index=False)
        pivot.to_excel(writer, sheet_name='Summary', index=False)

        for sheet_name in writer.sheets:
            ws_out = writer.sheets[sheet_name]
            for cell in ws_out[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1F4E79')
            for col in ws_out.columns:
                width = max(len(str(c.value or '')) for c in col)
                ws_out.column_dimensions[col[0].column_letter].width = width + 4

export_report(df, 'final_report.xlsx')

# ── BATCH PROCESSING ─────────────────────────────────────────
from pathlib import Path

# process all xlsx files in a folder
results = []
for file in Path('data/').glob('*.xlsx'):
    df_tmp = pd.read_excel(file)
    df_tmp['source'] = file.name           # track which file it came from
    results.append(df_tmp)

df_combined = pd.concat(results, ignore_index=True)
df_combined.to_excel('combined.xlsx', index=False)

# ── SCHEDULING ───────────────────────────────────────────────
# run a task on a schedule (pip install schedule)
# import schedule, time
#
# def daily_report():
#     df = pd.read_csv('live_data.csv')
#     export_report(df, f'report_{pd.Timestamp.now():%Y%m%d}.xlsx')
#
# schedule.every().day.at('08:00').do(daily_report)
# schedule.every(10).minutes.do(daily_report)
#
# while True:
#     schedule.run_pending()
#     time.sleep(60)